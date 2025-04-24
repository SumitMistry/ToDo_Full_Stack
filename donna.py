import os
import pdfplumber
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
import datetime

def read_config(config_file):
    """Read the configuration file for folder path and keywords."""
    with open(config_file, 'r') as f:
        lines = f.readlines()

    # Extract folder path and keywords
    folder_path = lines[0].strip().split('=')[1]
    keywords = lines[1].strip().split('=')[1].split(', ')

    return folder_path, keywords

def search_keyword_in_docx(doc_path, keywords):
    """Extract text from .docx files and search for multiple keywords."""
    doc = Document(doc_path)
    results = {keyword: [] for keyword in keywords}  # Store results for each keyword
    keyword_count = {keyword: 0 for keyword in keywords}  # Initialize keyword counts

    # Loop through paragraphs to find the keywords
    for i, paragraph in enumerate(doc.paragraphs):
        for keyword in keywords:
            paragraph_count = paragraph.text.lower().count(keyword.lower())
            if paragraph_count > 0:
                keyword_count[keyword] += paragraph_count
                results[keyword].append([keyword, keyword_count[keyword], i + 1, paragraph.text])

    return results, keyword_count

def search_keyword_in_pdf(pdf_path, keywords):
    """Extract text from a PDF file and search for multiple keywords using pdfplumber."""
    results = {keyword: [] for keyword in keywords}  # Store results for each keyword
    keyword_count = {keyword: 0 for keyword in keywords}  # Initialize keyword counts

    with pdfplumber.open(pdf_path) as pdf:
        # Loop through each page of the PDF
        for page_num, page in enumerate(pdf.pages):
            text = page.extract_text()

            # Search for keywords in the extracted text
            for keyword in keywords:
                paragraph_count = text.lower().count(keyword.lower())
                if paragraph_count > 0:
                    keyword_count[keyword] += paragraph_count
                    results[keyword].append([keyword, keyword_count[keyword], page_num + 1, text])

    return results, keyword_count

def save_to_excel(results, keyword_count, output_excel):
    """Save the results into an Excel file using openpyxl, with formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Keyword Search Results"

    # Add headers dynamically for each keyword
    headers = ['Doc Name', 'Keyword', 'Keyword Count', 'Paragraph Number', 'Paragraph Text']
    ws.append(headers)

    # Formatting for headers (bold and left-aligned)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = cell.alignment.copy(horizontal='left')

    # Loop through results for each keyword and add rows to Excel
    for doc_name, doc_results in results.items():
        for row in doc_results:
            ws.append([doc_name] + row)  # Add document name in the first column

    # Format columns to fit the content
    column_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            try:
                column_widths[cell.column] = max(column_widths.get(cell.column, 0), len(str(cell.value)))
            except:
                pass
    for col, width in column_widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width + 2

    # Add borders to all cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Save the workbook
    wb.save(output_excel)

def process_documents_in_folder(folder_path, keywords, output_excel):
    """Process all Word and PDF documents in a folder and save results to Excel."""
    results = {}  # Store results for each document
    keyword_count = {}  # Store keyword counts for all documents

    # Loop through all files in the folder
    for filename in os.listdir(folder_path):
        doc_path = os.path.join(folder_path, filename)

        if filename.lower().endswith(('.docx', '.docm')):  # Only process Word files
            print(f"Opening Word document: {filename}")
            doc_results, doc_keyword_count = search_keyword_in_docx(doc_path, keywords)
        elif filename.lower().endswith('.pdf'):  # Process PDF files
            print(f"Opening PDF document: {filename}")
            doc_results, doc_keyword_count = search_keyword_in_pdf(doc_path, keywords)
        else:
            continue  # Skip unsupported file types

        # Store results with the document name as the key
        results[filename] = []
        for keyword, keyword_results in doc_results.items():
            results[filename].extend(keyword_results)

        # Update overall keyword count
        keyword_count.update(doc_keyword_count)
        print(f"Processed document: {filename}")

    # Save results to Excel
    save_to_excel(results, keyword_count, output_excel)
    print(f"Search results have been saved to {output_excel}")

def main():
    # Determine the current running file's path and use the same location for config.txt
    current_dir = os.path.dirname(os.path.abspath(__file__))
    config_file = os.path.join(current_dir, 'config.txt')  # Full path to the config.txt file

    # Read the configuration
    folder_path, keywords = read_config(config_file)

    # Strip any leading/trailing spaces from the folder_path to avoid path issues
    folder_path = folder_path.strip()

    # Add timestamp to the output file name and save in the same folder
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_excel = os.path.join(folder_path, f'output_{timestamp}.xlsx')  # File name with timestamp suffix

    # Process all documents in the folder and save the results
    process_documents_in_folder(folder_path, keywords, output_excel)

if __name__ == '__main__':
    main()
